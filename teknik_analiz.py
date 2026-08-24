from binance.client import Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

client = Client()

def calculate_rsi(closes, period=14):
    if len(closes) < period + 1:
        return 0
    deltas = [closes[i] - closes[i-1] for i in range(1, len(closes))]
    gains = [d if d > 0 else 0 for d in deltas]
    losses = [-d if d < 0 else 0 for d in deltas]
    avg_gain = sum(gains[-period:]) / period
    avg_loss = sum(losses[-period:]) / period
    if avg_loss == 0:
        return 100
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

def calculate_macd(closes):
    if len(closes) < 26:
        return 0, 0, 0
    ema12 = sum(closes[-12:]) / 12
    ema26 = sum(closes[-26:]) / 26
    macd = ema12 - ema26
    signal = (macd + macd) / 2
    histogram = macd - signal
    return round(macd, 4), round(signal, 4), round(histogram, 4)

symbols = ['BTCUSDT', 'ETHUSDT', 'XRPUSDT', 'BNBUSDT', 'ADAUSDT']
print("Teknik analiz cekiliyor...\n")

wb = Workbook()
ws = wb.active
ws.title = "Analiz"
ws['A1'] = "Binance Futures - Teknik Analiz"
ws['A1'].font = Font(size=12, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws.merge_cells('A1:J1')

ws['A2'] = "Guncelleme: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
ws['A2'].font = Font(italic=True, size=9)

headers = ['Sembol', 'Fiyat', 'Degisim %', 'En Yuksek', 'En Dusuk', 'RSI(14)', 'SMA20', 'SMA50', 'MACD', 'Signal']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col)
    cell.value = h
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

row = 5
for symbol in symbols:
    try:
        ticker = client.futures_symbol_ticker(symbol=symbol)
        klines = client.futures_klines(symbol=symbol, interval='1h', limit=50)
        
        price = float(ticker.get('price', 0))
        change = float(ticker.get('priceChangePercent', 0))
        high = float(ticker.get('highPrice', 0))
        low = float(ticker.get('lowPrice', 0))
        
        closes = [float(k[4]) for k in klines]
        rsi = calculate_rsi(closes, 14)
        sma20 = sum(closes[-20:]) / 20 if len(closes) >= 20 else 0
        sma50 = sum(closes[-50:]) / 50 if len(closes) >= 50 else 0
        macd, signal, histogram = calculate_macd(closes)
        
        ws.cell(row=row, column=1).value = symbol
        ws.cell(row=row, column=2).value = round(price, 2)
        ws.cell(row=row, column=3).value = round(change, 2)
        ws.cell(row=row, column=4).value = round(high, 2)
        ws.cell(row=row, column=5).value = round(low, 2)
        ws.cell(row=row, column=6).value = round(rsi, 2)
        ws.cell(row=row, column=7).value = round(sma20, 2)
        ws.cell(row=row, column=8).value = round(sma50, 2)
        ws.cell(row=row, column=9).value = macd
        ws.cell(row=row, column=10).value = signal
        
        if change > 0:
            ws.cell(row=row, column=3).font = Font(color="00B050", bold=True)
        else:
            ws.cell(row=row, column=3).font = Font(color="FF0000", bold=True)
        
        if rsi > 70:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif rsi < 30:
            ws.cell(row=row, column=6).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        print("[OK] " + symbol + " - RSI: " + str(round(rsi, 2)) + " SMA20: " + str(round(sma20, 2)))
        row += 1
    except Exception as e:
        print("[HATA] " + symbol + ": " + str(e))

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws.column_dimensions[col].width = 14

wb.save('binance_teknik_analiz.xlsx')
print("\nBitti! binance_teknik_analiz.xlsx olusturuldu!")
print("Dosya: C:\\Users\\USER\\Desktop\\binance_teknik_analiz.xlsx")
