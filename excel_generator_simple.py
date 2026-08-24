from binance.client import Client
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import time

# Binance API (API Key istemiyoruz - public data)
client = Client()

def get_crypto_data(symbols):
    """Binance'den canlı veri çek"""
    data = []
    
    for symbol in symbols:
        try:
            # Canlı fiyat
            ticker = client.futures_symbol_ticker(symbol=symbol)
            
            # Mum grafiği verileri (son 50 mum)
            klines = client.futures_klines(symbol=symbol, interval='1h', limit=50)
            
            # Teknik göstergeler basit hesapla
            closes = [float(k[4]) for k in klines]
            
            # RSI hesapla
            rsi = calculate_rsi(closes, 14)
            
            # Moving Averages
            sma20 = sum(closes[-20:]) / 20
            sma50 = sum(closes[-50:]) / 50 if len(closes) >= 50 else sum(closes) / len(closes)
            
            data.append({
                'Sembol': symbol,
                'Fiyat': float(ticker['lastPrice']),
                '24s Değişim %': float(ticker['priceChangePercent']),
                '24s En Yüksek': float(ticker['highPrice']),
                '24s En Düşük': float(ticker['lowPrice']),
                'Hacim': float(ticker['volume']),
                'RSI (14)': round(rsi, 2),
                'SMA 20': round(sma20, 2),
                'SMA 50': round(sma50, 2),
            })
            
            print(f"✅ {symbol} veri çekildi")
            
        except Exception as e:
            print(f"❌ {symbol} hatası: {e}")
    
    return data

def calculate_rsi(closes, period=14):
    """RSI göstergesi hesapla"""
    if len(closes) < period + 1:
        return 0
    
    deltas = [closes[i] - closes[i-1] for i in range(1, len(closes))]
    gains = [d if d > 0 else 0 for d in deltas]
    losses = [-d if d < 0 else 0 for d in deltas]
    
    avg_gain = sum(gains[-period:]) / period
    avg_loss = sum(losses[-period:]) / period
    
    if avg_loss == 0:
        return 100
    
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return rsi

def create_excel(data, filename='crypto_analysis.xlsx'):
    """Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analiz"
    
    # Başlık
    ws['A1'] = "🚀 Binance Futures - Canlı Teknik Analiz"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:I1')
    
    # Güncelleme tarihi
    ws['A2'] = f"Güncelleme: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    
    # Sütun başlıkları
    headers = ['Sembol', 'Fiyat ($)', '24s Değişim (%)', '24s En Yüksek', 
               '24s En Düşük', 'Hacim', 'RSI (14)', 'SMA 20', 'SMA 50']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Veri satırları
    for row_num, row_data in enumerate(data, 5):
        ws.cell(row=row_num, column=1).value = row_data['Sembol']
        ws.cell(row=row_num, column=2).value = round(row_data['Fiyat'], 2)
        ws.cell(row=row_num, column=3).value = round(row_data['24s Değişim %'], 2)
        ws.cell(row=row_num, column=4).value = round(row_data['24s En Yüksek'], 2)
        ws.cell(row=row_num, column=5).value = round(row_data['24s En Düşük'], 2)
        ws.cell(row=row_num, column=6).value = round(row_data['Hacim'], 2)
        ws.cell(row=row_num, column=7).value = row_data['RSI (14)']
        ws.cell(row=row_num, column=8).value = row_data['SMA 20']
        ws.cell(row=row_num, column=9).value = row_data['SMA 50']
        
        # Renk kodlaması (Değişim %)
        change = row_data['24s Değişim %']
        if change > 0:
            ws.cell(row=row_num, column=3).font = Font(color="00B050")  # Yeşil
        else:
            ws.cell(row=row_num, column=3).font = Font(color="FF0000")  # Kırmızı
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 15
    
    # Kaydet
    wb.save(filename)
    print(f"\n✅ Excel başarıyla oluşturuldu: {filename}")
    return filename

# Ana program
if __name__ == "__main__":
    # Kripto paralar
    symbols = ['BTCUSDT', 'ETHUSDT', 'XRPUSDT', 'BNBUSDT', 'ADAUSDT', 'DOGEUSDT']
    
    print("📊 Binance Futures Teknik Analiz Excel Oluşturucu\n")
    print(f"📥 {', '.join(symbols)} için veri çekiliyor...\n")
    
    # Veri çek
    data = get_crypto_data(symbols)
    
    # Excel oluştur
    if data:
        create_excel(data, 'binance_teknik_analiz.xlsx')
        print("\n🎉 Tamamlandı! 'binance_teknik_analiz.xlsx' dosyasını açabilirsin.")
